import tkinter as tk
from PIL import Image, ImageTk
from tkinter import filedialog
import time
import os
import cv2 
from mtcnn import MTCNN
from keras.models import  load_model
from matplotlib.pylab import plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

lst_result = ['BINH','HUY','NAM','QUANG','TU','TUNG']
models = load_model(r'E:\hoctap\AI\project\dino\XLA\F.h5')

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GIAO DIEN")
        self.geometry("675x800") #kích thước tkinter

        # Tải ảnh từ tệp
        self.image = Image.open("nen2.jpg")
        self.image = self.image.resize((675, 750))
        # Chuyển đổi ảnh sang định dạng PhotoImage
        self.photo_nen = ImageTk.PhotoImage(self.image)
        # Tạo một widget Label và đưa ảnh lên đó
        self.label_nen = tk.Label(self, image=self.photo_nen)
        self.label_nen.pack()

        self.tieu_de = tk.Label(self, text="TRƯỜNG ĐẠI HỌC SƯ PHẠM KỸ THUẬT TP.HỒ CHÍ MINH \nKHOA CƠ KHÍ - CHẾ TẠO MÁY\nNGÀNH CƠ ĐIỆN TỬ \nBỘ MÔN MACHINE VISION",bg="white", fg="blue",)
        self.tieu_de.config(highlightthickness=0)
        self.tieu_de.config(font=("Arial", 12, "bold"))
        self.tieu_de.place(x=120, y=50)

        self.de_tai = tk.Label(self, text="ĐỀ TÀI: \nNHẬN DẠNG, ĐỊNH DANH VÀ ĐIỂM DANH SINH VIÊN",bg="white", fg="red")
        self.de_tai.config(highlightthickness=0)
        self.de_tai.config(font=("Arial", 14))
        self.de_tai.place(x=100, y=200)

        self.khoa = Image.open('khoa.png')
        self.khoa = self.khoa.resize((100,100))
        self.photo_khoa = ImageTk.PhotoImage(self.khoa)
        self.label_khoa = tk.Label(self, image=self.photo_khoa)
        self.label_khoa.config(highlightthickness=0)
        self.label_khoa.place(x=520, y=100)

        self.truong = Image.open('truong.jpg')
        self.truong = self.truong.resize((100,100))
        self.photo_truong = ImageTk.PhotoImage(self.truong)
        self.label_truong = tk.Label(self, image=self.photo_truong)
        self.label_truong.config(highlightthickness=0)
        self.label_truong.place(x=50, y=100)

        self.hinh = Image.open('mat.jpg')
        self.hinh = self.hinh.resize((150,156))
        self.photo_hinh = ImageTk.PhotoImage(self.hinh)
        self.label_hinh = tk.Label(self, image=self.photo_hinh)
        self.label_hinh.config(highlightthickness=0)
        self.label_hinh.place(x=250, y=250)

        # Tạo button để khởi chạy
        self.start_button = tk.Button(self, text="Start App", command=self.hide_shapes)
        self.start_button.place(x=300, y=420)

        # Tạo một đối tượng StringVar để lưu trữ giá trị trong ô văn bản
        self.GV = tk.Label(self, text="GVHD: TS. Nguyễn Văn Thái ",bg="white") 
        self.GV.config(highlightthickness=0)
        self.GV.config(font=("Arial", 12))
        self.GV.place(x=200, y=480)

        self.lop = tk.Label(self, text="Lớp: MAVI332529_22_2_03 - Nhóm:03",bg="white")
        self.lop.config(highlightthickness=0)
        self.lop.config(font=("Arial", 12))
        self.lop.place(x=200, y=510)

        self.SV = tk.Label(self, text="SVTH: nhóm 17 ",bg="white") 
        self.SV.config(highlightthickness=0)
        self.SV.config(font=("Arial", 12))
        self.SV.place(x=200, y=540)

        self.NAM = tk.Label(self, text="1. Đoàn Thanh Nam 20146506",bg="white") 
        self.NAM.config(highlightthickness=0)
        self.NAM.config(font=("Arial", 12))
        self.NAM.place(x=220, y=570)

        self.BINH = tk.Label(self, text="2. Vũ Đức Bình 20146478",bg="white")
        self.BINH.config(highlightthickness=0)
        self.BINH.config(font=("Arial", 12))
        self.BINH.place(x=220, y=600)

        self.QUANG = tk.Label(self, text="3. Lê Văn Quang 20146161",bg="white")
        self.QUANG.config(highlightthickness=0)
        self.QUANG.config(font=("Arial", 12))
        self.QUANG.place(x=220, y=630)

        self.Date = tk.Label(self, text="Thành phố Hồ Chí Minh - Tháng 5/2023",bg="white")
        self.Date.config(highlightthickness=0)
        self.Date.config(font=("Arial", 12))
        self.Date.place(x=190, y=700)

    def hide_shapes(self):
        # Ẩn các hình bằng phương thức pack_forget()
        self.Date.place_forget()
        self.QUANG.place_forget()
        self.NAM.place_forget()
        self.BINH.place_forget()
        self.SV.place_forget()
        self.lop.place_forget()
        self.GV.place_forget()
        self.label_hinh.place_forget()
        self.label_truong.place_forget()
        self.label_khoa.place_forget()
        self.label_khoa.place_forget()
        self.de_tai.place_forget()
        self.label_nen.pack_forget()
        self.tieu_de.place_forget()
        self.start_button.place_forget()
        time.sleep(1)
        self.title("CHƯƠNG TRÌNH CHÍNH")
        self.geometry("1000x1000") #kích thước tkinter
        self.configure(bg='light cyan')
        self.button = tk.Button(self, text="Upload",width=10, height=5, command=self.start)
        self.button.place(x=20, y=20)


    def start(self):
        self.file_path = filedialog.askopenfilename(title="Select file", filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif")])
        if self.file_path:
            self.test = Image.open(self.file_path)
            self.test = self.test.resize((400, 400))
            self.photo_test = ImageTk.PhotoImage(self.test)
            self.label_test = tk.Label(self, image=self.photo_test)
            self.label_test.config(highlightthickness=0)
            self.label_test.place(x=300, y=10)
            self.ten = tk.Label(self, text="HÌNH CẦN ĐIỂM DANH",bg="white") 
            self.ten.config(highlightthickness=0)
            self.ten.config(font=("Arial", 12))
            self.ten.place(x=415, y=420)
            self.predict_button = tk.Button(self, text="Predict",width=10, height=5, command=self.predict)
            self.predict_button.place(x=20, y=150)

    def predict(self):
        self.count=0
        self.filename=self.file_path
        #dung thu vien MTCNN
        self.detector=MTCNN()
        self.img1=plt.imread(self.filename)  
        self.faces=self.detector.detect_faces(self.img1)

        #tao folder chua anh snippets
        self.output_folder = "snippet_images"
        os.makedirs(self.output_folder, exist_ok=True)

        #chi hien thi cac hinh chu nhat ma co mat chu chua cat duoc mat
        #xoa tat ca cac anh truoc khi bam show tranh viec bi tran
        for self.file_name_snippet in os.listdir(self.output_folder):
            self.file_path = os.path.join(self.output_folder, self.file_name_snippet)
            if os.path.isfile(self.file_path):
                os.remove(self.file_path)

        os.makedirs(self.output_folder, exist_ok=True)

        # Iterate over the detected face bounding boxes
        for self.i, self.bbox in enumerate(self.faces):
            
            self.x, self.y, self.w, self.h = self.bbox['box']  # Extract the bounding box coordinates
            #predict model
            self.face_snippet = self.img1[self.y:self.y+self.h, self.x:self.x+self.w]
            self.resized_image = cv2.resize(self.face_snippet, (128,128))
            self.resized_image1 = self.resized_image.astype('float')*1./255
            self.resized_image1 = np.expand_dims(self.resized_image1, axis=0)
            self.predict = models.predict(self.resized_image1) 
            self.name=lst_result[np.argmax(self.predict)] 
            # Extract the face snippet from the original image
            
            self.save_path = os.path.join(self.output_folder, f"{self.name}.jpg")

            # Save the face snippet as an image
            
            cv2.imwrite(self.save_path, self.resized_image)
        self.text = tk.Label(self, text="DONE",fg="red")
        self.text.config(highlightthickness=0)
        self.text.config(font=("Arial", 12))
        self.text.place(x=22, y=250)
        self.show_button = tk.Button(self, text="SHOW",width=10, height=5, command=self.show)
        self.show_button.place(x=20, y=300)

    def show(self):
        self.dir_path = r"E:\hoctap\AI\project\dino\XLA\snippet_images"
        self.image_list = []
        for self.filename in os.listdir(self.dir_path):
            if self.filename.endswith(".jpg") or self.filename.endswith(".png"):
                self.image_path = os.path.join(self.dir_path, self.filename)
                self.image_list.append(self.image_path)

        self.x=20
        for self.image_path in self.image_list:
            self.img2 = Image.open(self.image_path)
            self.img2 = self.img2.resize((128,128))
            self.photo1 = ImageTk.PhotoImage(self.img2)
            self.label0 = tk.Label(self, image=self.photo1)
            self.label0.image = self.photo1  # keep a reference to the image to avoid garbage collection
            self.label0.place(x=self.x, y=500)
            self.file_name = os.path.basename(self.image_path)
            self.name = tk.Label(self, text=self.file_name,bg="white")
            self.name.config(font=("Arial", 10))
            self.name.place(x=self.x, y=650)
            self.x+=200
        
        self.sl = tk.Label(self, text="SỐ LƯỢNG SINH VIÊN LÀ: ",bg="white")
        self.sl.config(font=("Arial", 25))
        self.sl.place(x=200, y=700)
        self.sl = tk.Label(self, text=self.i+1,bg="white")
        self.sl.config(font=("Arial", 25))
        self.sl.place(x=660, y=700)

        self.in4_button = tk.Button(self, text="MORE INFO",width=20, height=5, bg="dark blue", fg="white", command=self.info)
        self.in4_button.config(font=("Arial", 14))
        self.in4_button.place(x=750, y=200)

    def info(self):
        self.destroy()
        window2 = tk.Tk()
        window2.title("THÔNG TIN")
        window2.geometry("1000x1000")
        window2.configure(bg='light yellow')

        self.y = 50
        for self.image_path in self.image_list:
            self.img2 = Image.open(self.image_path)
            self.img2 = self.img2.resize((128,128))
            self.photo1 = ImageTk.PhotoImage(self.img2)
            self.label0 = tk.Label(window2, image=self.photo1)
            self.label0.image = self.photo1  # keep a reference to the image to avoid garbage collection
            self.file_name = os.path.basename(self.image_path)
            self.wb = load_workbook(filename='info.xlsx', read_only=True)
            self.ws = self.wb.active
            if self.file_name == "BINH.jpg":
                self.value1 = self.ws['A1'].value
                self.value2 = self.ws['A2'].value
                self.value3 = self.ws['A3'].value
                self.text = f"{self.value1}\n{self.value2}\n{self.value3}"
                self.name = tk.Label(window2, text=self.text, bg="white")
                self.name.place(x=300, y=self.y)
            elif self.file_name == "HUY.jpg":
                self.value1 = self.ws['A4'].value
                self.value2 = self.ws['A5'].value
                self.value3 = self.ws['A6'].value
                self.text = f"{self.value1}\n{self.value2}\n{self.value3}"
                self.name = tk.Label(window2, text=self.text, bg="white")
                self.name.place(x=300, y=self.y)
            elif self.file_name == "QUANG.jpg":
                self.value1 = self.ws['A10'].value
                self.value2 = self.ws['A11'].value
                self.value3 = self.ws['A12'].value
                self.text = f"{self.value1}\n{self.value2}\n{self.value3}"
                self.name = tk.Label(window2, text=self.text, bg="white")
                self.name.place(x=300, y=self.y)
            elif self.file_name == "NAM.jpg":
                self.value1 = self.ws['A7'].value
                self.value2 = self.ws['A8'].value
                self.value3 = self.ws['A9'].value
                self.text = f"{self.value1}\n{self.value2}\n{self.value3}"
                self.name = tk.Label(window2, text=self.text, bg="white")
                self.name.place(x=300, y=self.y)
            elif self.file_name == "TU.jpg":
                self.value1 = self.ws['A13'].value
                self.value2 = self.ws['A14'].value
                self.value3 = self.ws['A15'].value
                self.text = f"{self.value1}\n{self.value2}\n{self.value3}"
                self.name = tk.Label(window2, text=self.text, bg="white")
                self.name.place(x=300, y=self.y)
            elif self.file_name == "TUNG.jpg":
                self.value1 = self.ws['A16'].value
                self.value2 = self.ws['A17'].value
                self.value3 = self.ws['A18'].value
                self.text = f"{self.value1}\n{self.value2}\n{self.value3}"
                self.name = tk.Label(window2, text=self.text, bg="white")
                self.name.place(x=300, y=self.y)
            self.name.config(font=("Arial", 10))           
            self.label0.place(x=20, y=self.y)
            self.y+=150


        self.sl = tk.Label(window2, text="SỐ LƯỢNG SINH VIÊN LÀ: ",bg="white")
        self.sl.config(font=("Arial", 25))
        self.sl.place(x=200, y=700)
        self.sl = tk.Label(window2, text=self.i+1,bg="white")
        self.sl.config(font=("Arial", 25))
        self.sl.place(x=660, y=700)
        self.check_button = tk.Button(window2, text="ĐIỂM DANH",width=20, height=5, bg="blue", fg="white", command=self.check)
        self.check_button.config(font=("Arial", 14))
        self.check_button.place(x=750, y=200)
        
        window2.mainloop()

    def check(self):
        self.wb = load_workbook("info.xlsx")
        self.ws = self.wb["Sheet1"]
        self.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.font = Font(color='00FF00')
        self.date = "E"                                     #điểm danh tuần 1 là B, tuần 2 là C,...
        for self.image_path in self.image_list:
            self.file_name = os.path.basename(self.image_path)
            if self.file_name == "BINH.jpg":
                self.text = f"{self.date}1"
                self.ws[self.text] = 'yes'
                self.ws[self.text].fill = self.fill
                self.ws[self.text].font = self.font
            elif self.file_name == "HUY.jpg":
                self.text = f"{self.date}4"
                self.ws[self.text] = 'yes'
                self.ws[self.text].fill = self.fill
                self.ws[self.text].font = self.font
            elif self.file_name == "QUANG.jpg":
                self.text = f"{self.date}10"
                self.ws[self.text] = 'yes'
                self.ws[self.text].fill = self.fill
                self.ws[self.text].font = self.font
            elif self.file_name == "NAM.jpg":
                self.text = f"{self.date}7"
                self.ws[self.text] = 'yes'
                self.ws[self.text].fill = self.fill
                self.ws[self.text].font = self.font
            elif self.file_name == "TU.jpg":
                self.text = f"{self.date}13"
                self.ws[self.text] = 'yes'
                self.ws[self.text].fill = self.fill
                self.ws[self.text].font = self.font
            elif self.file_name == "TUNG.jpg":
                self.text = f"{self.date}16"
                self.ws[self.text] = 'yes'
                self.ws[self.text].fill = self.fill
                self.ws[self.text].font = self.font
            
        self.wb.save("info.xlsx")






app=App()
app.mainloop()